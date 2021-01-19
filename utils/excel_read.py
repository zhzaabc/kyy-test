import openpyxl


class ParseExcel():
    def __init__(self,excel_path,sheetName):
        self.wb = openpyxl.load_workbook(excel_path)
        self.sheet = self.wb[sheetName]
        self.active = self.wb.active
        self.min_row = self.active.min_row #最小行
        self.max_row = self.active.max_row #最大行
        self.min_col = self.active.min_column #最大列
        self.max_col = self.active.max_column #最大列


    def getDataFromSheet(self):
        datalist = []
        for i in range(self.min_row+1,self.max_row+1):#从行开始遍历
            row_info=[]
            for j in range(self.min_col,self.max_col+1):
                row_info.append(self.sheet.cell(i, j).value)
            datalist.append(row_info)  # 注意，python以对齐来确定循环的所定义区域
        return datalist

# if  __name__ == '__main__':
#     excel_path = './../data/login_excel.xlsx'
#     sheetName = 'user'
#     parse = ParseExcel(excel_path,sheetName)
#     print(parse.getDataFromSheet()[0][0])